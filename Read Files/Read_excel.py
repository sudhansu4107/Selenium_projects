import openpyxl

Workbook=openpyxl.load_workbook('E:\\Automation Project\\Files\\Test1_Excel.xlsx')
Sheet=Workbook['Sheet1']
rows=Sheet.max_row
cols=Sheet.max_column
print(f'The total number of  rows={rows},columns={cols}')

# Print all  the content present in the Excel file
for r in range(1,rows+1):
    for  c in range(1,cols+1):
        print(Sheet.cell(row=r,column=c).value,end=" ")
    print()