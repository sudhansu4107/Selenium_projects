import openpyxl


class Excel_data:
    rows = ''
    cols = ''
    Product_name_list = []
    Qty = []

    def __init__(self):
        self.file = openpyxl.load_workbook('G:\\Selenium\\testdata\\Item_excel.xlsx')
        self.sheet = self.file['Sheet1']
        Excel_data.rows = self.sheet.max_row
        Excel_data.cols = self.sheet.max_column
        print(Excel_data.rows, Excel_data.cols)

    # def input(self):
    #
    #     self.Fetch_Productname()
    #     self.Fetch_Qty()

    def Fetch_Productname(self):
        for r in range(2, Excel_data.rows + 1):
            if self.sheet.cell(row=r, column=1).value is not None:
                print(self.sheet.cell(row=r, column=1).value)
                Excel_data.Product_name_list.append(self.sheet.cell(row=r, column=1).value)
        print(f'Fetch_Product_name process  completed')
        return Excel_data.Product_name_list

    def Fetch_Qty(self):
        for r in range(2, Excel_data.rows + 1):
            if self.sheet.cell(row=r, column=2).value is not None:
                print(self.sheet.cell(row=r, column=2).value)
                Excel_data.Qty.append(self.sheet.cell(row=r, column=2).value)
        print(f'Fetch_Qty  process  completed.')
        return Excel_data.Qty

    def Update_Item_price(self, Item_price):
        for r in range(2, len(Item_price) + 2):
            self.sheet.cell(row=r, column=3).value = str(Item_price[r - 2])
            print(f'Item price updated')
            self.file.save('G:\\Selenium\\testdata\\Item_excel.xlsx')

    def Update_Total_amt(self, Total_amt):
        for r in range(2,len(Total_amt)+2):
            self.sheet.cell(row=r, column=4).value = str(Total_amt[r-2])
            print(f'Total amt price updated.')
            self.file.save('G:\\Selenium\\testdata\\Item_excel.xlsx')

# obj = Excel_data()
# product_list = obj.Fetch_Productname()
# Qty_list = obj.Fetch_Qty()
# print(f'The product_list={product_list},The Qty is ;{Qty_list}')
