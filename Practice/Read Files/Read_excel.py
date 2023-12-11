import openpyxl

Workbook = openpyxl.load_workbook('/Practice/Files\\Test1_Excel.xlsx')
Sheet = Workbook['Sheet1']
rows = Sheet.max_row
cols = Sheet.max_column
print(f'The total number of  rows={rows},columns={cols}')

# # Print all  the content present in the Excel file
# for r in range(1, rows + 1):
#     for c in range(1, cols + 1):
#         print(Sheet.cell(row=r, column=c).value, end=" ")
#     print()

# Write into  Excel file
Test_Data = ['Yogesh', 'Yogesh005@gmail.com','Yogesh012', 'Male', 'Fresher']

Updated_Rows = len(Test_Data)

# One Test data has been added
rows = rows + 1

Text_value = ''


def Update_Excel():
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            if Sheet.cell(row=r, column=c).value is None:
                Sheet.cell(row=r, column=c).value = Test_Data[c-1]


Update_Excel()
